import base64
import hashlib
import json
import os
import re
import zlib
from xml.dom.minidom import parseString
from xml.etree.ElementTree import ParseError
from xml.parsers.expat import ExpatError

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def choices_to_options(choice_value):
    """ CHOICE -> {value=xxx,label=xxx} """
    data = []
    for item in choice_value:
        key_value = dict(value=item[0], label=item[1])
        data.append(key_value)
    return data


def phone_check(phone):
    """ 手机check """
    phone_re = re.match("^0*(13|18|17|15|14|16|19)\d{9}$", str(phone))
    if not phone_re:
        return False
    return True


def email_check(email):
    """ 邮箱 check """
    email_re = re.match(
        "^[A-Za-z0-9\u4e00-\u9fa5_-]+@[a-zA-Z0-9_-]+(\.[a-zA-Z0-9_-]+)+$", email
    )
    if not email_re:
        return False
    return True


def read_json_file(filename):
    """
    读取.json文件 返回dict
    :param filename: 文件名
    :return:
    """
    if not os.path.isfile(filename):
        raise FileExistsError
    with open(filename, "r") as config_file:
        config_data = json.load(config_file)
        return config_data


def iflytek_sso_response_parse(content):
    """
    解析iflytek sso verify result
    :param content:
    :return:
    {
        'res': 0,
        'msg': '',
        'data': {
            'username': 'sss'  // 域账号
            'first_name': '中文名'
        }
    }
    """
    try:
        dom = parseString(content)
        if dom.documentElement.nodeName != "cas:serviceResponse":
            return dict(res=1, msg="no serviceResponse field", data={})
        failures = dom.getElementsByTagName("cas:authenticationFailure")
        if len(failures) > 0:
            return dict(res=1, msg=failures[0].firstChild.data, dadta={})
        successes = dom.getElementsByTagName("cas:authenticationSuccess")
        if len(successes) > 0:
            users = dom.getElementsByTagName("cas:user")
            username = str(users[0].firstChild.data)
            first_name = dom.getElementsByTagName("cas:userName")[0].firstChild.data
            return dict(
                res=0, data=dict(username=username, first_name=first_name), msg=""
            )
        return dict(res=0, msg="parse error1")
    except (ExpatError, ParseError, KeyError):
        return dict(res=1, msg="parse raise error")


def check_int(value):
    try:
        int(value)
        return True
    except ValueError:
        return False


def match_ip(value):
    """
    过滤出传入值的ip
    :param value:  str -sdfsd-10.103.15.256
    :return:  str 10.103.15.256
    """
    pattern = re.compile(
        r"(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)\.(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)\.(25[0-5]|2[0-4]"
        r"\d|[0-1]\d{2}|[1-9]?\d)\.(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)"
    )
    data = pattern.search(value)
    if data is None:
        return ""
    if len(value) > data.regs[-1][-1]:
        last_value = value[data.regs[-1][-1]]
        if check_int(last_value):
            return ""
    return data.group()


def match_url(value):
    """
    过滤出传入值的web url
    :param value:  str -sdfs-https://www.baidu.com
    :return:  str https://www.baidu.com
    """

    pattern = re.compile(r"((https|http|ftp|rtsp|mms)?:\/\/)[^\s]+")
    data = pattern.search(value)
    if data is None:
        return ""
    return data.group()


def data_for_excel(datas):
    """
    转化为 excel适用
    :param datas:
            [{'key1': 'value1', 'key2': 'value2'},{'key1': 'value3', 'key2': 'value4'}]
    :return: [['key1', 'key2'], ['value1', 'value2'], ['value3', 'value4']]
    """
    if not datas:
        return [[]]
    datas_excel = []
    headers = datas[0].keys()
    datas_excel.append(list(headers))
    for x in datas:
        item = []
        for y in headers:
            item.append(x[y])
        datas_excel.append((item))
    return datas_excel


def excel_write(dest_filename, datas, sheet_name=""):
    """
    生成excel文件
    :param dest_filename:  string 文件名
    :param datas:
        * [['This is A1', 'This is B1', 'This is C1'], ]
        * **or** [{'A' : 'This is A1', 'C' : 'This is C1'}]
        * **or** [{1 : 'This is A1', 3 : 'This is C1'}]
    :param sheet_name:  sheet_name
    :return:
    """
    wb = Workbook()
    ws1 = wb.active
    if not sheet_name:
        ws1.title = "Sheet1"
    else:
        ws1.title = sheet_name
    for data in datas:
        ws1.append(data)
    wb.save(filename=dest_filename)
    return dest_filename


def read_excel(filename):
    """
    读取excel文件 返回所有数据
    :param dest_filename:
    :return: datas  []
    """
    wb = load_workbook(filename=filename)
    datas = []
    for item in wb.active.rows:
        datas.append([x.internal_value for x in item])
    return datas


def hashcrc32_dict(data):
    """
    给json对象求hash值
    :param data:
    :return:
    """
    return zlib.crc32(json.dumps(data, sort_keys=True).encode("utf-8"))


def is_sub_dict(d, sub_d):
    for key, value in sub_d.items():
        if key in d and d.get(key) == value:
            pass
        else:
            return False
    return True
